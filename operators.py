import bpy
import os
import subprocess
from . import geometry_utils, boundary_conditions, therm_export, therm_import, therm_runner
import xml.etree.ElementTree as ET
import shutil
import os


# Import funkcji do obliczania grubości
from .geometry_utils import (
    calculate_material_thickness, 
    get_mesh_dimensions, 
    calculate_smart_thickness,
    get_curve_points_world
)

# Operatory dla sprawdzania i naprawy geometrii
class THERM_OT_check_normals(bpy.types.Operator):
    """Sprawdza czy wszystkie faces mają normalne skierowane w górę"""
    bl_idname = "therm.check_normals"
    bl_label = "Sprawdź normalne"
    bl_description = "Sprawdza czy wszystkie faces mają normalne skierowane w górę"
    
    def execute(self, context):
        result = geometry_utils.check_and_fix_normals()
        if result == {'FINISHED'}:
            self.report({'INFO'}, "Wszystkie normalne są skierowane w górę")
        return result

class THERM_OT_clean_to_boundary(bpy.types.Operator):
    """Czyści geometrię do samej obwiedni"""
    bl_idname = "therm.clean_to_boundary"
    bl_label = "Wyczyść do obwiedni"
    bl_options = {'REGISTER', 'UNDO'}
    
    def execute(self, context):
        try:
            if bpy.context.mode != 'EDIT_MESH':
                bpy.ops.object.mode_set(mode='EDIT')
            
            bpy.ops.mesh.select_all(action='SELECT')
            bpy.ops.mesh.region_to_loop()
            bpy.ops.mesh.select_all(action='INVERT')
            bpy.ops.mesh.delete(type='EDGE_FACE')
            bpy.ops.mesh.select_all(action='SELECT')
            bpy.ops.mesh.edge_face_add()
            bpy.ops.object.mode_set(mode='OBJECT')
            self.report({'INFO'}, "Wyczyścino geometrię do obwiedni")
            return {'FINISHED'}
        except Exception as e:
            self.report({'ERROR'}, f"Błąd: {str(e)}")
            return {'CANCELLED'}

class THERM_OT_check_vertices(bpy.types.Operator):
    """Sprawdza czy wierzchołki mają współrzędne z dokładnością do 0.1m"""
    bl_idname = "therm.check_vertices"
    bl_label = "Sprawdź wierzchołki"
    bl_description = "Sprawdza czy wierzchołki mają współrzędne z dokładnością do 0.1m"
    
    def execute(self, context):
        result = geometry_utils.check_and_round_vertices()
        if result == {'FINISHED'}:
            self.report({'INFO'}, "Wszystkie wierzchołki są wielokrotnościami 0.1m")
        return result

class THERM_OT_round_vertices(bpy.types.Operator):
    """Zaokrągla wierzchołki zaznaczonych obiektów"""
    bl_idname = "therm.round_vertices"
    bl_label = "Zaokrągl wierzchołki"
    bl_description = "Zaokrągla współrzędne wierzchołków zaznaczonych obiektów do określonej precyzji"
    
    def execute(self, context):
        selected_objects = [obj for obj in context.selected_objects if obj.type == 'MESH']
        
        if not selected_objects:
            self.report({'WARNING'}, "Zaznacz przynajmniej jeden obiekt siatki")
            return {'CANCELLED'}
        
        precision = float(context.scene.therm_props.round_precision)
        vertices_modified = geometry_utils.round_vertices_to_precision(precision)
        
        self.report({'INFO'}, f"Zaokrąglono {vertices_modified} wierzchołków z precyzją {precision}m")
        return {'FINISHED'}

class THERM_OT_force_round_vertices(bpy.types.Operator):
    """Wymusza zaokrąglenie wszystkich wierzchołków zaznaczonych obiektów"""
    bl_idname = "therm.force_round_vertices"
    bl_label = "Wymuś zaokrąglenie wierzchołków"
    bl_description = "Wymusza zaokrąglenie wszystkich wierzchołków bez sprawdzania"
    
    def execute(self, context):
        selected_objects = [obj for obj in context.selected_objects if obj.type == 'MESH']
        
        if not selected_objects:
            self.report({'WARNING'}, "Zaznacz przynajmniej jeden obiekt siatki")
            return {'CANCELLED'}
        
        precision = float(context.scene.therm_props.round_precision)
        vertices_modified = geometry_utils.round_vertices_to_precision(precision)
        
        self.report({'INFO'}, f"Zaokrąglono {vertices_modified} wierzchołków z precyzją {precision}m")
        return {'FINISHED'}

# Operatory potwierdzające
class THERM_OT_recalc_normals_confirm(bpy.types.Operator):
    """Operator potwierdzający odwrócenie normalnych"""
    bl_idname = "therm.recalc_normals_confirm"
    bl_label = "Odwróć normalne"
    bl_description = "Odwróć normalne faces do góry"
    
    def execute(self, context):
        geometry_utils.recalc_normals_upward()
        self.report({'INFO'}, "Odwrócono normalne do góry")
        return {'FINISHED'}
    
    def invoke(self, context, event):
        return context.window_manager.invoke_confirm(self, event)

class THERM_OT_round_vertices_confirm(bpy.types.Operator):
    """Operator potwierdzający zaokrąglenie wierzchołków"""
    bl_idname = "therm.round_vertices_confirm"
    bl_label = "Zaokrągl wierzchołki"
    bl_description = "Zaokrągl wierzchołki do dokładności 0.1m"
    
    def execute(self, context):
        precision = 0.1
        vertices_modified = geometry_utils.round_vertices_to_precision(precision)
        self.report({'INFO'}, f"Zaokrąglono {vertices_modified} wierzchołków do {precision}m")
        return {'FINISHED'}
    
    def invoke(self, context, event):
        return context.window_manager.invoke_confirm(self, event)

# Operatory do odwracania faces skierowanych w dół
class THERM_OT_flip_downward_faces(bpy.types.Operator):
    """Odwraca tylko faces skierowane w dół"""
    bl_idname = "therm.flip_downward_faces"
    bl_label = "Odwróć faces skierowane w dół"
    bl_description = "Automatycznie znajduje i odwraca tylko faces które są skierowane w dół"
    bl_options = {'REGISTER', 'UNDO'}
    
    def execute(self, context):
        threshold = context.scene.therm_props.flip_threshold
        faces_flipped = geometry_utils.flip_downward_faces_with_threshold(threshold)
        self.report({'INFO'}, f"Odwrócono {faces_flipped} faces skierowanych w dół (próg: {threshold})")
        return {'FINISHED'}

class THERM_OT_quick_flip_downward_faces(bpy.types.Operator):
    """Szybkie odwracanie faces skierowanych w dół"""
    bl_idname = "therm.quick_flip_downward_faces"
    bl_label = "Szybko odwróć faces w dół"
    bl_description = "Szybkie odwracanie faces skierowanych w dół z domyślnym progiem"
    
    def execute(self, context):
        faces_flipped = geometry_utils.flip_downward_faces_only()
        self.report({'INFO'}, f"Odwrócono {faces_flipped} faces skierowanych w dół")
        return {'FINISHED'}

# Operatory dla tworzenia krzywych ręcznie
class THERM_OT_create_ti_edges(bpy.types.Operator):
    """Tworzy krzywe Ti ze wszystkich zaznaczonych krawędzi"""
    bl_idname = "therm.create_ti_edges"
    bl_label = "Utwórz krzywe Ti"
    
    def execute(self, context):
        curve_objs = boundary_conditions.create_continuous_curve_from_edges('Ti')
        if curve_objs:
            coll_name = boundary_conditions.get_therm_collection_name('Ti')
            self.report({'INFO'}, f"Utworzono {len(curve_objs)} krzywych Ti w kolekcji '{coll_name}'")
        else:
            self.report({'WARNING'}, "Nie zaznaczono żadnych krawędzi")
        return {'FINISHED'}

class THERM_OT_create_te_edges(bpy.types.Operator):
    """Tworzy krzywe Te ze wszystkich zaznaczonych krawędzi"""
    bl_idname = "therm.create_te_edges"
    bl_label = "Utwórz krzywe Te"
    
    def execute(self, context):
        curve_objs = boundary_conditions.create_continuous_curve_from_edges('Te')
        if curve_objs:
            coll_name = boundary_conditions.get_therm_collection_name('Te')
            self.report({'INFO'}, f"Utworzono {len(curve_objs)} krzywych Te w kolekcji '{coll_name}'")
        else:
            self.report({'WARNING'}, "Nie zaznaczono żadnych krawędzi")
        return {'FINISHED'}

class THERM_OT_create_adiabatic_edges(bpy.types.Operator):
    """Tworzy krzywe Adiabatic ze wszystkich zaznaczonych krawędzi"""
    bl_idname = "therm.create_adiabatic_edges"
    bl_label = "Utwórz krzywe Adiabatic"
    
    def execute(self, context):
        curve_objs = boundary_conditions.create_continuous_curve_from_edges('Adiabatic')
        if curve_objs:
            coll_name = boundary_conditions.get_therm_collection_name('Adiabatic')
            self.report({'INFO'}, f"Utworzono {len(curve_objs)} krzywych Adiabatic w kolekcji '{coll_name}'")
        else:
            self.report({'WARNING'}, "Nie zaznaczono żadnych krawędzi")
        return {'FINISHED'}

class THERM_OT_create_ufactor_edges(bpy.types.Operator):
    """Tworzy krzywe U-Factor ze wszystkich zaznaczonych krawędzi"""
    bl_idname = "therm.create_ufactor_edges"
    bl_label = "Utwórz krzywe U-Factor"
    
    def execute(self, context):
        ufactor_name = context.scene.therm_edge_props.ufactor_name
        if not ufactor_name:
            self.report({'WARNING'}, "Ustaw nazwę U-Factor")
            return {'CANCELLED'}
        
        curve_objs = boundary_conditions.create_continuous_curve_from_edges('UFactor', ufactor_name)
        if curve_objs:
            coll_name = boundary_conditions.get_therm_collection_name('UFactor', ufactor_name)
            self.report({'INFO'}, f"Utworzono {len(curve_objs)} krzywych U-Factor w kolekcji '{coll_name}'")
        else:
            self.report({'WARNING'}, "Nie zaznaczono żadnych krawędzi")
        return {'FINISHED'}

# Operatory dla automatycznego tworzenia krzywych
class THERM_OT_create_adiabatic_auto(bpy.types.Operator):
    """Automatycznie tworzy krzywe Adiabatic na zewnętrznych krawędziach z kierunkiem lewoskrętnym"""
    bl_idname = "therm.create_adiabatic_auto"
    bl_label = "Auto Adiabatic (zewnętrzne krawędzie)"
    bl_description = "Automatycznie znajduje zewnętrzne krawędzie i tworzy krzywe Adiabatic z kierunkiem lewoskrętnym"
    
    def execute(self, context):
        created_curves = boundary_conditions.create_auto_curves_on_external_edges('Adiabatic')
        
        if created_curves:
            self.report({'INFO'}, f"Utworzono {len(created_curves)} krzywych Adiabatic na zewnętrznych krawędziach")
        else:
            self.report({'WARNING'}, "Nie znaleziono zewnętrznych krawędzi")
        
        return {'FINISHED'}

class THERM_OT_create_ti_auto(bpy.types.Operator):
    """Automatycznie tworzy krzywe Ti na zewnętrznych krawędziach z kierunkiem lewoskrętnym"""
    bl_idname = "therm.create_ti_auto"
    bl_label = "Auto Ti (zewnętrzne krawędzie)"
    bl_description = "Automatycznie znajduje zewnętrzne krawędzie i tworzy krzywe Ti z kierunkiem lewoskrętnym"
    
    def execute(self, context):
        created_curves = boundary_conditions.create_auto_curves_on_external_edges('Ti')
        
        if created_curves:
            self.report({'INFO'}, f"Utworzono {len(created_curves)} krzywych Ti na zewnętrznych krawędziach")
        else:
            self.report({'WARNING'}, "Nie znaleziono zewnętrznych krawędzi")
        
        return {'FINISHED'}

class THERM_OT_create_te_auto(bpy.types.Operator):
    """Automatycznie tworzy krzywe Te na zewnętrznych krawędziach z kierunkiem lewoskrętnym"""
    bl_idname = "therm.create_te_auto"
    bl_label = "Auto Te (zewnętrzne krawędzie)"
    bl_description = "Automatycznie znajduje zewnętrzne krawędzie i tworzy krzywe Te z kierunkiem lewoskrętnym"
    
    def execute(self, context):
        created_curves = boundary_conditions.create_auto_curves_on_external_edges('Te')
        
        if created_curves:
            self.report({'INFO'}, f"Utworzono {len(created_curves)} krzywych Te na zewnętrznych krawędziach")
        else:
            self.report({'WARNING'}, "Nie znaleziono zewnętrznych krawędzi")
        
        return {'FINISHED'}

class THERM_OT_create_ufactor_auto(bpy.types.Operator):
    """Automatycznie tworzy krzywe U-Factor na zewnętrznych krawędziach z kierunkiem lewoskrętnym"""
    bl_idname = "therm.create_ufactor_auto"
    bl_label = "Auto U-Factor (zewnętrzne krawędzie)"
    bl_description = "Automatycznie znajduje zewnętrzne krawędzie i tworzy krzywe U-Factor z kierunkiem lewoskrętnym"
    
    def execute(self, context):
        ufactor_name = bpy.context.scene.therm_edge_props.ufactor_name
        if not ufactor_name:
            self.report({'WARNING'}, "Ustaw nazwę U-Factor")
            return {'CANCELLED'}
        
        created_curves = boundary_conditions.create_auto_curves_on_external_edges('UFactor', ufactor_name)
        
        if created_curves:
            self.report({'INFO'}, f"Utworzono {len(created_curves)} krzywych U-Factor na zewnętrznych krawędziach")
        else:
            self.report({'WARNING'}, "Nie znaleziono zewnętrznych krawędzi")
        
        return {'FINISHED'}

# Operatory dla uruchamiania THERM
class THERM_OT_run_therm_calculation_thmx(bpy.types.Operator):
    """Uruchom obliczenia THERM z plikiem .thmx"""
    bl_idname = "therm.run_therm_calculation_thmx"
    bl_label = "Uruchom obliczenia THERM (.thmx)"
    bl_description = "Uruchom obliczenia w THERM z plikiem .thmx (wymaga THERM7.exe)"
    
    def execute(self, context):
        runner = therm_runner.THERMRunner()
        result_type, message = runner.run_calculation_thmx(context)
        self.report(result_type, message)
        return {'FINISHED'}

class THERM_OT_run_therm_calculation_thm(bpy.types.Operator):
    """Uruchom obliczenia THERM z plikiem .thm"""
    bl_idname = "therm.run_therm_calculation_thm"
    bl_label = "Uruchom obliczenia THERM (.thm)"
    bl_description = "Uruchom obliczenia w THERM z plikiem .thm (wymaga THERM7.exe)"
    
    def execute(self, context):
        runner = therm_runner.THERMRunner()
        result_type, message = runner.run_calculation_thm(context)
        self.report(result_type, message)
        return {'FINISHED'}

class THERM_OT_open_therm_folder(bpy.types.Operator):
    """Otwórz folder z plikami THERM"""
    bl_idname = "therm.open_therm_folder"
    bl_label = "Otwórz folder THERM"
    bl_description = "Otwórz folder z plikami THERM"
    
    def execute(self, context):
        runner = therm_runner.THERMRunner()
        result_type, message = runner.open_therm_folder(context)
        self.report(result_type, message)
        return {'FINISHED'}

# Operator importu THERM
class THERM_OT_import_from_therm(bpy.types.Operator):
    """Importuj plik THERM (.thmx) do Blendera"""
    bl_idname = "therm.import_from_therm"
    bl_label = "Importuj z THERM"
    bl_description = "Importuj plik THERM (.thmx) do Blendera"
    
    filepath: bpy.props.StringProperty(subtype="FILE_PATH")
    
    def execute(self, context):
        if not self.filepath:
            self.report({'ERROR'}, "Nie wybrano pliku")
            return {'CANCELLED'}
        
        if not os.path.exists(self.filepath):
            self.report({'ERROR'}, f"Plik nie istnieje: {self.filepath}")
            return {'CANCELLED'}
        
        try:
            importer = therm_import.THERMImporter()
            success = importer.import_therm_file(self.filepath)
            if success:
                self.report({'INFO'}, f"Zaimportowano plik THERM: {os.path.basename(self.filepath)}")
            else:
                self.report({'ERROR'}, "Błąd importu pliku THERM")
        except Exception as e:
            self.report({'ERROR'}, f"Błąd importu: {str(e)}")
            return {'CANCELLED'}
        
        return {'FINISHED'}
    
    def invoke(self, context, event):
        context.window_manager.fileselect_add(self)
        return {'RUNNING_MODAL'}

# Operator eksportu THERM
class THERM_OT_export_to_therm(bpy.types.Operator):
    """Eksportuj do pliku THERM"""
    bl_idname = "therm.export_to_therm"
    bl_label = "Eksportuj do THERM"
    
    def execute(self, context):
        exporter = therm_export.THERMExporter()
        result_type, message = exporter.export_to_therm(context)
        
        if result_type == {'INFO'} and context.scene.therm_props.open_export_folder:
            self.open_export_folder()
        
        self.report(result_type, message)
        return {'FINISHED'}
    
    def open_export_folder(self):
        """Otwiera folder z wyeksportowanym plikiem"""
        blend_filepath = bpy.data.filepath
        if not blend_filepath:
            return
        
        folder_path = os.path.dirname(blend_filepath)
        try:
            if platform.system() == "Windows":
                os.startfile(folder_path)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", folder_path])
            else:
                subprocess.Popen(["xdg-open", folder_path])
        except Exception as e:
            print(f"Nie można otworzyć folderu: {e}")

# Klasa bazowa dla tworzenia sekcji U z grubościami materiałów
class THERM_OT_create_usection_base(bpy.types.Operator):
    """Klasa bazowa dla tworzenia sekcji U z grubościami materiałów"""
    bl_options = {'REGISTER', 'UNDO'}
    
    socket_map = {
        'u1': 'Socket_25',
        'usection': 'Socket_2',
        'y': 'Socket_24', 
        'u-value': 'Socket_26',
        'r01': 'Socket_28', 'r02': 'Socket_29', 'r03': 'Socket_30', 'r04': 'Socket_31',
        'r05': 'Socket_32', 'r06': 'Socket_33', 'r07': 'Socket_34', 'r08': 'Socket_35',
        'r09': 'Socket_36', 'r10': 'Socket_37',
        'ti': 'Socket_22', 'te': 'Socket_23',
        'm01': 'Socket_8', 'm02': 'Socket_14', 'm03': 'Socket_15', 'm04': 'Socket_16',
        'm05': 'Socket_17', 'm06': 'Socket_18', 'm07': 'Socket_19', 'm08': 'Socket_20',
        'm09': 'Socket_13', 'm10': 'Socket_12',
        # Dodaj sockety dla grubości jeśli są dostępne w Geometry Nodes
        't01': 'Socket_3', 't02': 'Socket_4', 't03': 'Socket_5', 't04': 'Socket_6',
        't05': 'Socket_7', 't06': 'Socket_9', 't07': 'Socket_10', 't08': 'Socket_11',
    }
    
    usection_name: bpy.props.StringProperty()
    
    def find_ti_curves_from_selected(self):
        """Znajduje krzywe Ti tylko z zaznaczonych obiektów/kolekcji"""
        ti_curves = []
        selected_objects = bpy.context.selected_objects
        
        # Sprawdź wszystkie zaznaczone krzywe w kolekcjach Ti
        for obj in selected_objects:
            if obj.type == 'CURVE':
                # Sprawdź czy obiekt jest w kolekcji Ti
                for coll in obj.users_collection:
                    if coll.name.startswith('THERM_Ti='):
                        ti_curves.append(obj)
                        break
        
        return ti_curves
    
    def ensure_usection_collection(self):
        """Tworzy kolekcję THERM_USections jeśli nie istnieje"""
        collection_name = "THERM_USections"
        if collection_name not in bpy.data.collections:
            coll = bpy.data.collections.new(collection_name)
            bpy.context.scene.collection.children.link(coll)
        return bpy.data.collections[collection_name]
    
    def check_usection_exists(self, usection_name):
        """Sprawdza czy krzywa USection już istnieje"""
        target_name = f"USection_{usection_name}"
        return target_name in bpy.data.objects
    
    def calculate_all_thicknesses(self, mesh_objects, ti_curve, te_curve):
        """Oblicza grubości dla wszystkich obiektów siatki"""
        thicknesses = []
        
        print("📐 INTELIGENTNE OBLICZANIE GRUBOŚCI MATERIAŁÓW:")
        print("═" * 60)
        
        for i, mesh_obj in enumerate(mesh_objects):
            # Użyj inteligentnej metody obliczania grubości
            final_thickness = calculate_smart_thickness(mesh_obj, ti_curve, te_curve)
            thicknesses.append(final_thickness)
        
        # Podsumowanie wszystkich grubości
        print("📊 PODSUMOWANIE GRUBOŚCI:")
        total_thickness = sum(thicknesses)
        print(f"📏 SUMA GRUBOŚCI WSZYSTKICH WARSTW: {total_thickness:.4f}m")
        
        # Sprawdź czy suma grubości jest realistyczna
        if total_thickness > 5.0:  # 5 metrów
            print("⚠️  UWAGA: Suma grubości przekracza 5m - sprawdź poprawność!")
        elif total_thickness < 0.1:  # 10 cm
            print("⚠️  UWAGA: Suma grubości mniejsza niż 10cm - sprawdź poprawność!")
        
        print("═" * 60)
        return thicknesses
    
    def set_basic_values(self, modifier, usection_name, ti_curve, te_curve):
        """Ustawia podstawowe wartości w Geometry Nodes"""
        available_inputs = list(modifier.keys())
        
        # Ustaw USection
        if 'usection' in self.socket_map and self.socket_map['usection'] in available_inputs:
            modifier[self.socket_map['usection']] = usection_name
            print(f"✅ Ustawiono {self.socket_map['usection']} (USection) = {usection_name}")
        
        # Ustaw Ti
        ti_curves = self.find_ti_curves_from_selected()
        if 'ti' in self.socket_map and self.socket_map['ti'] in available_inputs and ti_curves:
            modifier[self.socket_map['ti']] = ti_curves[0]
            print(f"✅ Ustawiono {self.socket_map['ti']} (Ti) = {ti_curves[0].name}")
        
        # Ustaw Te  
        te_curves = self.find_all_te_curves()
        if 'te' in self.socket_map and self.socket_map['te'] in available_inputs and te_curves:
            modifier[self.socket_map['te']] = te_curves[0] 
            print(f"✅ Ustawiono {self.socket_map['te']} (Te) = {te_curves[0].name}")
    
    def calculate_u_value(self, thicknesses, conductivities, ti_rsi, te_rse):
        """Oblicza współczynnik U (U-Value) na podstawie grubości, conductivity i oporów"""
        try:
            print("🧮 OBLICZANIE WSPÓŁCZYNNIKA U-VALUE:")
            print("═" * 50)
            
            # Opory przejmowania ciepła (domyślne wartości z właściwości sceny)
            Rsi = ti_rsi  # Opór wewnętrzny
            Rse = te_rse  # Opór zewnętrzny
            
            print(f"📊 Opory przejmowania ciepła:")
            print(f"   Rsi = {Rsi:.3f} m²K/W")
            print(f"   Rse = {Rse:.3f} m²K/W")
            
            # Oblicz opory dla każdej warstwy: R = d / λ
            layer_resistances = []
            total_material_resistance = 0.0
            
            print("\n📋 OBLICZANIE OPORÓW WARSTW:")
            print("No. | Grubość [m] | Conductivity [W/mK] | Opór R [m²K/W]")
            print("----|-------------|-------------------|----------------")
            
            for i, (thickness, conductivity) in enumerate(zip(thicknesses, conductivities)):
                if thickness > 0 and conductivity > 0:
                    layer_resistance = thickness / conductivity
                    layer_resistances.append(layer_resistance)
                    total_material_resistance += layer_resistance
                    print(f"{i+1:2d} | {thickness:11.4f} | {conductivity:17.4f} | {layer_resistance:14.4f}")
                else:
                    layer_resistances.append(0.0)
                    print(f"{i+1:2d} | {thickness:11.4f} | {conductivity:17.4f} | {'BŁĄD':>14}")
            
            # Oblicz całkowity opór cieplny RT
            RT = Rsi + total_material_resistance + Rse
            
            print(f"\n📊 SUMA OPORÓW MATERIAŁÓW: {total_material_resistance:.4f} m²K/W")
            print(f"📊 CAŁKOWITY OPÓR RT: {RT:.4f} m²K/W")
            print(f"   Rsi: {Rsi:.4f} m²K/W")
            print(f"   ΣR_materiałów: {total_material_resistance:.4f} m²K/W") 
            print(f"   Rse: {Rse:.4f} m²K/W")
            
            # Oblicz współczynnik U = 1 / RT
            if RT > 0:
                u_value = 1.0 / RT
                print(f"🎯 WSPÓŁCZYNNIK U-VALUE: {u_value:.6f} W/m²K")
            else:
                u_value = 0.0
                print(f"❌ BŁĄD: Całkowity opór RT = 0")
            
            print("═" * 50)
            return u_value, RT, total_material_resistance
            
        except Exception as e:
            print(f"❌ Błąd obliczania U-Value: {e}")
            return 0.0, 0.0, 0.0

    def set_geometry_nodes_values_with_thickness(self, curve_obj, usection_name, modifier, mesh_objects, ti_curve, te_curve):
        """Ustawia wartości w geometry nodes z grubościami materiałów i oblicza U-Value"""
        try:
            available_inputs = list(modifier.keys())
            
            print(f"🎯 Ustawianie Geometry Nodes dla {curve_obj.name} z grubościami i U-Value:")
            print("═" * 50)
            
            # Oblicz grubości materiałów
            thicknesses = self.calculate_all_thicknesses(mesh_objects, ti_curve, te_curve)
            
            # Pobierz conductivities
            conductivities = []
            for mesh_obj in mesh_objects:
                conductivity = self.get_material_conductivity(mesh_obj)
                conductivities.append(conductivity)
            
            # Pobierz wartości Rsi i Rse z właściwości sceny
            ti_rsi = bpy.context.scene.therm_edge_props.ti_rsi
            te_rse = bpy.context.scene.therm_edge_props.te_rse
            
            # Oblicz U-Value
            u_value, total_resistance, material_resistance = self.calculate_u_value(
                thicknesses, conductivities, ti_rsi, te_rse
            )
            
            # Ustaw podstawowe wartości
            self.set_basic_values(modifier, usection_name, ti_curve, te_curve)
            
            # Ustaw U-Value w Geometry Nodes jeśli socket istnieje
            if 'u-value' in self.socket_map and self.socket_map['u-value'] in available_inputs:
                try:
                    modifier[self.socket_map['u-value']] = u_value
                    print(f"✅ Ustawiono {self.socket_map['u-value']} (U-Value) = {u_value:.6f} W/m²K")
                except Exception as e:
                    print(f"❌ Błąd ustawiania U-Value: {e}")
            
            # Ustaw obiekty, conductivities i grubości
            objects_set = 0
            conductivities_set = 0
            thicknesses_set = 0
            
            # Mapowanie socketów
            object_keys = ['m01', 'm02', 'm03', 'm04', 'm05', 'm06', 'm07', 'm08', 'm09', 'm10']
            conductivity_keys = ['r01', 'r02', 'r03', 'r04', 'r05', 'r06', 'r07', 'r08', 'r09', 'r10']
            thickness_keys = ['t01', 't02', 't03', 't04', 't05', 't06', 't07', 't08', 't09', 't10']
            
            for i, (obj_key, cond_key, thick_key) in enumerate(zip(object_keys, conductivity_keys, thickness_keys)):
                if i >= len(mesh_objects):
                    break
                
                # Ustaw obiekt
                if obj_key in self.socket_map and self.socket_map[obj_key] in available_inputs:
                    modifier[self.socket_map[obj_key]] = mesh_objects[i]
                    print(f"✅ Ustawiono {self.socket_map[obj_key]} ({obj_key.upper()}) = {mesh_objects[i].name}")
                    objects_set += 1
                
                # Ustaw conductivity
                if cond_key in self.socket_map and self.socket_map[cond_key] in available_inputs:
                    try:
                        modifier[self.socket_map[cond_key]] = conductivities[i]
                        print(f"✅ Ustawiono {self.socket_map[cond_key]} ({cond_key.upper()}) = {conductivities[i]:.4f} W/mK")
                        conductivities_set += 1
                    except Exception as e:
                        print(f"❌ Błąd ustawiania conductivity: {e}")
                
                # Ustaw grubość
                if thick_key in self.socket_map and self.socket_map[thick_key] in available_inputs:
                    try:
                        modifier[self.socket_map[thick_key]] = thicknesses[i]
                        print(f"✅ Ustawiono {self.socket_map[thick_key]} ({thick_key.upper()}) = {thicknesses[i]:.4f} m")
                        thicknesses_set += 1
                    except Exception as e:
                        print(f"❌ Błąd ustawiania grubości: {e}")
            
            # Podsumowanie
            print("═" * 50)
            print(f"📊 PODSUMOWANIE:")
            print(f"   🏗️  Obiekty: {objects_set}/{len(mesh_objects)}")
            print(f"   🔥 Conductivities: {conductivities_set}/{len(mesh_objects)}") 
            print(f"   📏 Grubości: {thicknesses_set}/{len(mesh_objects)}")
            print(f"   🎯 U-Value: {u_value:.6f} W/m²K")
            print(f"   📊 Całkowity opór RT: {total_resistance:.4f} m²K/W")
            
            # Wypisz tabelę materiałów
            print("\n📋 TABELA MATERIAŁÓW:")
            print("No. | Obiekt | Grubość [m] | Conductivity [W/mK] | Opór R [m²K/W]")
            print("----|--------|-------------|-------------------|----------------")
            for i, (mesh_obj, thickness, conductivity) in enumerate(zip(mesh_objects, thicknesses, conductivities)):
                if thickness > 0 and conductivity > 0:
                    layer_resistance = thickness / conductivity
                    print(f"{i+1:2d} | {mesh_obj.name:6} | {thickness:11.4f} | {conductivity:17.4f} | {layer_resistance:14.4f}")
                else:
                    print(f"{i+1:2d} | {mesh_obj.name:6} | {thickness:11.4f} | {conductivity:17.4f} | {'BŁĄD':>14}")
            
            return True
                        
        except Exception as e:
            print(f"❌ Błąd ustawiania geometry nodes: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def create_usection_geometry_nodes(self, curve_obj, usection_name):
        """Dodaje geometry nodes do krzywej U-Section i ustawia wartości z grubościami"""
        try:
            # Usuń istniejące modyfikatory geometry nodes
            for mod in curve_obj.modifiers:
                if mod.type == 'NODES':
                    curve_obj.modifiers.remove(mod)
            
            node_group_name = "THERM U-Section"
            if node_group_name not in bpy.data.node_groups:
                self.create_usection_node_group(node_group_name)
            
            node_group = bpy.data.node_groups[node_group_name]
            modifier = curve_obj.modifiers.new(name=f"THERM_U_{usection_name}", type='NODES')
            modifier.node_group = node_group
            
            # Znajdź potrzebne obiekty
            mesh_objects = [obj for obj in bpy.context.selected_objects if obj.type == 'MESH']
            ti_curves = self.find_ti_curves_from_selected()
            te_curves = self.find_all_te_curves()
            
            if not ti_curves or not te_curves:
                print("❌ Brak krzywych Ti lub Te")
                return False, "Brak krzywych Ti lub Te"
            
            # Ustaw wartości z grubościami
            success = self.set_geometry_nodes_values_with_thickness(
                curve_obj, usection_name, modifier, mesh_objects, ti_curves[0], te_curves[0]
            )
            
            return success, "Success"
            
        except Exception as e:
            print(f"❌ Błąd dodawania geometry nodes: {e}")
            return False, str(e)
    
    def create_usection_node_group(self, node_group_name):
        """Tworzy grupę geometry nodes dla U-Section jeśli nie istnieje"""
        if node_group_name in bpy.data.node_groups:
            return bpy.data.node_groups[node_group_name]
        
        node_group = bpy.data.node_groups.new(node_group_name, 'GeometryNodeTree')
        
        # Dodaj inputy
        group_input = node_group.nodes.new('NodeGroupInput')
        group_input.location = (-400, 0)
        
        # Dodaj outputy  
        group_output = node_group.nodes.new('NodeGroupOutput')
        group_output.location = (400, 0)
        
        # Zdefiniuj socket-y
        node_group.interface.new_socket('Geometry', in_out='INPUT', socket_type='NodeSocketGeometry')
        node_group.interface.new_socket('Ti', in_out='INPUT', socket_type='NodeSocketGeometry')
        node_group.interface.new_socket('Te', in_out='INPUT', socket_type='NodeSocketGeometry')
        node_group.interface.new_socket('Object', in_out='INPUT', socket_type='NodeSocketGeometry')
        node_group.interface.new_socket('USection', in_out='INPUT', socket_type='NodeSocketString')
        node_group.interface.new_socket('Geometry', in_out='OUTPUT', socket_type='NodeSocketGeometry')
        
        return node_group
    
    def get_material_conductivity(self, mesh_object):
        """Pobiera wartość conductivity z materiału obiektu - POPRAWIONA WERSJA"""
        try:
            if not mesh_object.data.materials:
                print(f"    ❌ Obiekt {mesh_object.name} nie ma materiałów")
                return 0.04  # Wartość domyślna
            
            # Weź pierwszy materiał
            material = mesh_object.data.materials[0]
            if not material:
                return 0.04
            
            print(f"    🔍 Szukam conductivity w materiale: {material.name}")
            
            # METODA 1: Sprawdź CUSTOM PROPERTIES (najpierw)
            if 'conductivity' in material:
                conductivity = material['conductivity']
                print(f"    ✅ Znaleziono conductivity w custom properties: {conductivity}")
                return conductivity
            if 'thermal_conductivity' in material:
                conductivity = material['thermal_conductivity']
                print(f"    ✅ Znaleziono thermal_conductivity w custom properties: {conductivity}")
                return conductivity
            if 'lambda' in material:
                conductivity = material['lambda']
                print(f"    ✅ Znaleziono lambda w custom properties: {conductivity}")
                return conductivity
            
            # METODA 2: Sprawdź NODE'y MATERIALU (główna metoda)
            if material.use_nodes:
                print(f"    🔍 Przeszukuję node'y materiału...")
                
                for node in material.node_tree.nodes:
                    node_name = node.name
                    node_label = getattr(node, 'label', '') or ''
                    
                    print(f"    🔍 Node: {node_name}, Label: '{node_label}'")
                    
                    # SPRAWDŹ CZY NODE MA LABEL "conductivity"
                    if node_label and 'conductivity' in node_label.lower():
                        print(f"    ✅ Znaleziono node z label 'conductivity': {node_name}")
                        
                        # Sprawdź wartość w outputach node'a
                        if hasattr(node, 'outputs') and node.outputs:
                            for output in node.outputs:
                                if hasattr(output, 'default_value'):
                                    conductivity = output.default_value
                                    print(f"    ✅ Pobrano conductivity z output {output.name}: {conductivity}")
                                    return conductivity
                        
                        # Jeśli nie ma outputów, sprawdź czy to node Value
                        if hasattr(node, 'inputs') and node.inputs and hasattr(node.inputs[0], 'default_value'):
                            conductivity = node.inputs[0].default_value
                            print(f"    ✅ Pobrano conductivity z node Value: {conductivity}")
                            return conductivity
                    
                    # SPRAWDŹ CZY NODE MA NAME "conductivity"
                    if 'conductivity' in node_name.lower():
                        print(f"    ✅ Znaleziono node z name 'conductivity': {node_name}")
                        
                        if hasattr(node, 'outputs') and node.outputs:
                            for output in node.outputs:
                                if hasattr(output, 'default_value'):
                                    conductivity = output.default_value
                                    print(f"    ✅ Pobrano conductivity z output: {conductivity}")
                                    return conductivity
                        
                        if hasattr(node, 'inputs') and node.inputs and hasattr(node.inputs[0], 'default_value'):
                            conductivity = node.inputs[0].default_value
                            print(f"    ✅ Pobrano conductivity z node inputs: {conductivity}")
                            return conductivity
                    
                    # SPRAWDŹ SPECJALNIE DLA NODE'ÓW VALUE
                    if node.bl_idname == 'ShaderNodeValue':
                        node_value = getattr(node, 'outputs', [])
                        if node_value and hasattr(node_value[0], 'default_value'):
                            conductivity = node_value[0].default_value
                            # Sprawdź czy wartość jest realistyczna (nie 0.0)
                            if conductivity > 0.001:
                                print(f"    ✅ Pobrano conductivity z ShaderNodeValue: {conductivity}")
                                return conductivity
            
            # METODA 3: Sprawdź po nazwie materiału (fallback)
            material_name_lower = material.name.lower()
            print(f"    🔍 Sprawdzam nazwę materiału: {material_name_lower}")
            
            material_conductivity_map = {
                'beton': 1.7, 'concrete': 1.7, 'cement': 1.7,
                'cegła': 0.8, 'brick': 0.8, 'ceramika': 0.8,
                'drewno': 0.15, 'wood': 0.15, 'timber': 0.15,
                'szkło': 1.0, 'glass': 1.0,
                'stal': 50.0, 'steel': 50.0, 'metal': 50.0,
                'aluminium': 200.0, 'aluminum': 200.0,
                'izolacja': 0.04, 'insulation': 0.04, 'wełna': 0.04, 'wool': 0.04,
                'styropian': 0.035, 'eps': 0.035, 'xps': 0.035,
                'l0_80': 0.80, 'l0_04': 0.04, 'l0_15': 0.15, 'l0_035': 0.035,
                'l0_113': 0.113, 'l0_113_rama': 0.113
            }
            
            for material_keyword, conductivity_value in material_conductivity_map.items():
                if material_keyword in material_name_lower:
                    print(f"    ✅ Znaleziono conductivity po nazwie '{material_keyword}': {conductivity_value}")
                    return conductivity_value
            
            print(f"    ⚠️  Nie znaleziono conductivity, używam wartości domyślnej 0.04 W/mK")
            return 0.04  # Wartość domyślna dla izolacji
                
        except Exception as e:
            print(f"    ❌ Błąd pobierania conductivity: {e}")
            import traceback
            traceback.print_exc()
            return 0.04  # Wartość domyślna
        
    def find_all_te_curves(self):
        """Znajduje wszystkie krzywe Te w scenie"""
        te_curves = []
        for coll in bpy.data.collections:
            if coll.name.startswith('THERM_Te='):
                for obj in coll.objects:
                    if obj.type == 'CURVE':
                        te_curves.append(obj)
        return te_curves
    
    def create_usection(self, usection_name):
        """Główna funkcja tworząca sekcję U tylko z zaznaczonych obiektów"""
        try:
            target_name = f"USection_{usection_name}"
            
            # Sprawdź czy krzywa już istnieje
            if self.check_usection_exists(usection_name):
                self.report({'WARNING'}, f"Krzywa {target_name} już istnieje!")
                return {'CANCELLED'}
            
            # Znajdź krzywe Ti tylko z zaznaczonych obiektów
            ti_curves = self.find_ti_curves_from_selected()
            if not ti_curves:
                self.report({'WARNING'}, "Nie znaleziono zaznaczonych krzywych Ti")
                return {'CANCELLED'}
            
            print(f"Znaleziono {len(ti_curves)} zaznaczonych krzywych Ti")
            
            # Używamy tylko pierwszej zaznaczonej krzywej Ti
            ti_curve = ti_curves[0]
            
            print(f"Kopiowanie krzywej: {ti_curve.name} -> {target_name}")
            
            # Sprawdź ponownie czy krzywa już nie powstała (dla bezpieczeństwa)
            if self.check_usection_exists(usection_name):
                self.report({'WARNING'}, f"Krzywa {target_name} już istnieje! (ponowne sprawdzenie)")
                return {'CANCELLED'}
            
            # Utwórz kolekcję
            usection_coll = self.ensure_usection_collection()
            
            # Skopiuj krzywą
            new_curve = ti_curve.copy()
            new_curve.data = ti_curve.data.copy()
            new_curve.name = target_name
            new_curve.data.name = target_name
            
            # Dodaj do kolekcji
            usection_coll.objects.link(new_curve)
            
            # Odśwież scenę
            bpy.context.view_layer.update()
            
            # Sprawdź czy krzywa została poprawnie utworzona
            if new_curve.name not in bpy.data.objects:
                self.report({'ERROR'}, f"Błąd: Krzywa {target_name} nie została utworzona")
                return {'CANCELLED'}
            
            print(f"✅ Krzywa {target_name} została utworzona")
            
            # Dodaj geometry nodes
            success, message = self.create_usection_geometry_nodes(new_curve, usection_name)
            if success:
                print(f"✅ Geometry Nodes dodane do {target_name}")
                self.report({'INFO'}, f"Utworzono krzywą {target_name} z Geometry Nodes")
            else:
                print(f"❌ Błąd Geometry Nodes: {message}")
                self.report({'WARNING'}, f"Utworzono krzywą {target_name} ale błąd Geometry Nodes: {message}")
            
            # Odznacz wszystko i zaznacz nową krzywą
            bpy.ops.object.select_all(action='DESELECT')
            new_curve.select_set(True)
            bpy.context.view_layer.objects.active = new_curve
            
            return {'FINISHED'}
            
        except Exception as e:
            self.report({'ERROR'}, f"Błąd tworzenia U{usection_name}: {str(e)}")
            import traceback
            traceback.print_exc()
            return {'CANCELLED'}

# Operatory dla tworzenia poszczególnych sekcji U
class THERM_OT_create_usection_1(THERM_OT_create_usection_base):
    """Tworzy sekcję U1 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_1"
    bl_label = "Utwórz U1"
    
    def execute(self, context):
        return self.create_usection("U1")

class THERM_OT_create_usection_2(THERM_OT_create_usection_base):
    """Tworzy sekcję U2 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_2"
    bl_label = "Utwórz U2"
    
    def execute(self, context):
        return self.create_usection("U2")

class THERM_OT_create_usection_3(THERM_OT_create_usection_base):
    """Tworzy sekcję U3 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_3"
    bl_label = "Utwórz U3"
    
    def execute(self, context):
        return self.create_usection("U3")

class THERM_OT_create_usection_4(THERM_OT_create_usection_base):
    """Tworzy sekcję U4 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_4"
    bl_label = "Utwórz U4"
    
    def execute(self, context):
        return self.create_usection("U4")

class THERM_OT_create_usection_5(THERM_OT_create_usection_base):
    """Tworzy sekcję U5 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_5"
    bl_label = "Utwórz U5"
    
    def execute(self, context):
        return self.create_usection("U5")

class THERM_OT_create_usection_6(THERM_OT_create_usection_base):
    """Tworzy sekcję U6 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_6"
    bl_label = "Utwórz U6"
    
    def execute(self, context):
        return self.create_usection("U6")

class THERM_OT_create_usection_7(THERM_OT_create_usection_base):
    """Tworzy sekcję U7 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_7"
    bl_label = "Utwórz U7"
    
    def execute(self, context):
        return self.create_usection("U7")

class THERM_OT_create_usection_8(THERM_OT_create_usection_base):
    """Tworzy sekcję U8 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_8"
    bl_label = "Utwórz U8"
    
    def execute(self, context):
        return self.create_usection("U8")

class THERM_OT_create_usection_9(THERM_OT_create_usection_base):
    """Tworzy sekcję U9 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_9"
    bl_label = "Utwórz U9"
    
    def execute(self, context):
        return self.create_usection("U9")

class THERM_OT_create_usection_10(THERM_OT_create_usection_base):
    """Tworzy sekcję U10 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_10"
    bl_label = "Utwórz U10"
    
    def execute(self, context):
        return self.create_usection("U10")

class THERM_OT_create_usection_11(THERM_OT_create_usection_base):
    """Tworzy sekcję U11 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_11"
    bl_label = "Utwórz U11"
    
    def execute(self, context):
        return self.create_usection("U11")

class THERM_OT_create_usection_12(THERM_OT_create_usection_base):
    """Tworzy sekcję U12 z zaznaczonych krzywych Ti"""
    bl_idname = "therm.create_usection_12"
    bl_label = "Utwórz U12"
    
    def execute(self, context):
        return self.create_usection("U12")

# Operatory debugujące
class THERM_OT_debug_usections(bpy.types.Operator):
    """Debugowanie sekcji U"""
    bl_idname = "therm.debug_usections"
    bl_label = "Debug USections"
    
    def execute(self, context):
        print("=== DEBUG USECTIONS ===")
        
        # Sprawdź wszystkie krzywe
        all_curves = [obj for obj in bpy.data.objects if obj.type == 'CURVE']
        print(f"Wszystkie krzywe ({len(all_curves)}):")
        for curve in all_curves:
            print(f"  - {curve.name} (w kolekcjach: {[coll.name for coll in curve.users_collection]})")
        
        # Sprawdź kolekcje THERM_USections
        if "THERM_USections" in bpy.data.collections:
            usection_coll = bpy.data.collections["THERM_USections"]
            print(f"Kolekcja THERM_USections ({len(usection_coll.objects)} obiektów):")
            for obj in usection_coll.objects:
                print(f"  - {obj.name}")
        else:
            print("Kolekcja THERM_USections nie istnieje")
        
        # Sprawdź czy istnieją USection_*
        usection_objects = [obj for obj in bpy.data.objects if obj.name.startswith('USection_')]
        print(f"Obiekty USection_* ({len(usection_objects)}):")
        for obj in usection_objects:
            print(f"  - {obj.name}")
        
        print("=== KONIEC DEBUG ===")
        return {'FINISHED'}

class THERM_OT_debug_sockets(bpy.types.Operator):
    """Debugowanie socketów Geometry Nodes"""
    bl_idname = "therm.debug_sockets"
    bl_label = "Debug Sockets"
    
    def execute(self, context):
        print("=== DEBUG SOCKETS ===")
        
        # Sprawdź zaznaczone obiekty z Geometry Nodes
        selected_objects = bpy.context.selected_objects
        for obj in selected_objects:
            if obj.type == 'CURVE':
                for modifier in obj.modifiers:
                    if modifier.type == 'NODES' and modifier.node_group:
                        print(f"Obiekt: {obj.name}")
                        print(f"Grupa: {modifier.node_group.name}")
                        print("Socket-y INPUT:")
                        
                        # Wejścia grupy
                        for item in modifier.node_group.interface.items_tree:
                            if hasattr(item, 'in_out') and item.in_out == 'INPUT':
                                print(f"  - {item.name} (typ: {item.socket_type})")
                        
                        # Socket-y w modyfikatorze
                        available_inputs = list(modifier.keys())
                        print(f"  Dostępne w modyfikatorze: {sorted(available_inputs)}")
                        print("---")
        
        print("=== KONIEC DEBUG SOCKETS ===")
        return {'FINISHED'}



# USUŃ te importy:
# import pandas as pd

# ZASTĄP funkcje bez użycia pandas:

class THERM_OT_export_to_excel(bpy.types.Operator):
    """Eksportuj dane do pliku Excel na podstawie pliku .thmx i Geometry Nodes"""
    bl_idname = "therm.export_to_excel"
    bl_label = "Eksportuj do Excel"
    bl_description = "Eksportuj dane U-Value do pliku Excel na podstawie pliku .thmx i Geometry Nodes"
    
    def execute(self, context):
        try:
            blend_filepath = bpy.data.filepath
            if not blend_filepath:
                self.report({'ERROR'}, "Zapisz plik Blender przed eksportem do Excel")
                return {'CANCELLED'}
            
            # Pobierz nazwę pliku blendera
            blend_filename = os.path.splitext(os.path.basename(blend_filepath))[0]
            blend_directory = os.path.dirname(blend_filepath)
            
            # Ścieżki do plików
            thmx_file = os.path.join(blend_directory, f"{blend_filename}.thmx")
            excel_file = os.path.join(blend_directory, f"{blend_filename}.xlsx")
            template_file = os.path.join(blend_directory, "wzorzec.xlsx")
            
            # Sprawdź czy plik .thmx istnieje
            if not os.path.exists(thmx_file):
                self.report({'ERROR'}, f"Plik {thmx_file} nie istnieje. Najpierw wyeksportuj do THERM.")
                return {'CANCELLED'}
            
            # Sprawdź czy plik wzorca istnieje
            if not os.path.exists(template_file):
                self.report({'ERROR'}, f"Plik wzorca {template_file} nie istnieje.")
                return {'CANCELLED'}
            
            print(f"🔍 Przetwarzanie pliku: {thmx_file}")
            
            # Krok 1: Przetwórz plik .thmx - znajdź wartości PHI i strumień ciepła
            phi_data, phi_heat_flux = self.extract_phi_factors_from_thmx(thmx_file)  # ZWRACA TERAZ 2 WARTOŚCI
            temperatures = self.find_temperatures_from_thmx(thmx_file)
            
            # Krok 2: Pobierz wartości U z Geometry Nodes zaznaczonych krzywych
            u_values_from_gn = self.get_u_values_from_selected_curves()
            curve_lengths = self.get_lengths_from_selected_curves()
            
            # Krok 3: Skopiuj i wypełnij szablon Excel
            success = self.copy_and_fill_excel_template(
                template_file, excel_file, phi_data, temperatures, 
                u_values_from_gn, curve_lengths, thmx_file, phi_heat_flux  # DODANO phi_heat_flux
            )
            
            if success:
                self.report({'INFO'}, f"Utworzono plik Excel: {excel_file}")
                return {'FINISHED'}
            else:
                self.report({'ERROR'}, "Błąd podczas tworzenia pliku Excel")
                return {'CANCELLED'}
                
        except Exception as e:
            self.report({'ERROR'}, f"Błąd eksportu do Excel: {str(e)}")
            import traceback
            traceback.print_exc()
            return {'CANCELLED'}
    
    def extract_phi_factors_from_thmx(self, thmx_file):
        """Ekstrahuje wartości PHI z pliku .thmx i zwraca dane PHI oraz strumień ciepła"""
        try:
            with open(thmx_file, 'r', encoding='utf-8') as file:
                content = file.read()
            
            # Usuń namespace dla prostszego parsowania
            content = content.replace(' xmlns="http://windows.lbl.gov"', '')
            root = ET.fromstring(content)
            
            phi_data = {}
            phi_heat_flux = None
            
            # Szukaj wszystkich PHI-values
            for phi_elem in root.findall('.//PHI-value'):
                tag_elem = phi_elem.find('Tag')
                if tag_elem is not None:
                    tag = tag_elem.text
                    value_elem = phi_elem.find('Value')
                    if value_elem is not None:
                        value_str = value_elem.get('value')
                        if value_str and value_str != 'NA':
                            phi_value = float(value_str)
                            phi_data[tag] = phi_value
                            print(f"Znaleziono PHI {tag} = {phi_value}")
            
            # Szukaj strumienia ciepła dla PHI w U-factors
            for u_factor in root.findall('.//U-factors'):
                tag_elem = u_factor.find('Tag')
                if tag_elem is not None and 'PHI' in tag_elem.text:
                    print(f"Znaleziono U-factors dla {tag_elem.text}")
                    
                    # Pobierz DeltaT
                    delta_t_elem = u_factor.find('DeltaT')
                    delta_t = float(delta_t_elem.get('value')) if delta_t_elem is not None else 40.0
                    
                    # Szukaj projekcji z "Total length"
                    for projection in u_factor.findall('Projection'):
                        length_type_elem = projection.find('Length-type')
                        if length_type_elem is not None and 'Total length' in length_type_elem.text:
                            
                            # Pobierz długość
                            length_elem = projection.find('Length')
                            length_value = float(length_elem.get('value')) if length_elem is not None else 0.0
                            
                            # Pobierz wartość U
                            u_factor_elem = projection.find('U-factor')
                            if u_factor_elem is not None:
                                u_value_str = u_factor_elem.get('value')
                                if u_value_str and u_value_str != 'NA':
                                    u_value = float(u_value_str)
                                    
                                    # OBLICZENIE STRUMIENIA CIEPŁA (tak jak w Twoim skrypcie)
                                    heat_flux_per_meter = u_value * delta_t  # W/m
                                    total_heat_flux = heat_flux_per_meter * (length_value / 1000.0)  # zamiana mm na m
                                    
                                    phi_heat_flux = total_heat_flux
                                    print(f"🔥 Obliczono strumień ciepła dla PHI: {total_heat_flux:.6f} W")
                                    print(f"   U = {u_value}, ΔT = {delta_t}, L = {length_value}mm")
            
            # Jeśli nie znaleziono PHI, szukaj w starych PSI (dla kompatybilności)
            if not phi_data:
                print("Nie znaleziono PHI, szukam PSI...")
                for psi_elem in root.findall('.//PSI-value'):
                    tag_elem = psi_elem.find('Tag')
                    if tag_elem is not None:
                        tag = tag_elem.text
                        value_elem = psi_elem.find('Value')
                        if value_elem is not None:
                            value_str = value_elem.get('value')
                            if value_str and value_str != 'NA':
                                phi_value = float(value_str)
                                # Zamień PSI na PHI w nazwie
                                phi_tag = tag.replace('PSI', 'PHI') if 'PSI' in tag else tag
                                phi_data[phi_tag] = phi_value
                                print(f"Znaleziono PSI {tag} -> PHI {phi_tag} = {phi_value}")
            
            return phi_data, phi_heat_flux
            
        except Exception as e:
            print(f"Błąd ekstrakcji PHI z .thmx: {e}")
            return {}, None

    def extract_heat_flux_from_thmx(self, thmx_file):
        """Ekstrahuje strumień ciepła z pliku .thmx dla PHI"""
        try:
            with open(thmx_file, 'r', encoding='utf-8') as file:
                content = file.read()
            
            # Usuń namespace dla prostszego parsowania
            content = content.replace(' xmlns="http://windows.lbl.gov"', '')
            root = ET.fromstring(content)
            
            # Szukaj PHI w U-factors i oblicz strumień ciepła
            for u_factor in root.findall('.//U-factors'):
                tag_elem = u_factor.find('Tag')
                if tag_elem is not None and 'PHI' in tag_elem.text:
                    print(f"Znaleziono U-factors dla {tag_elem.text}")
                    
                    # Pobierz DeltaT
                    delta_t_elem = u_factor.find('DeltaT')
                    delta_t = float(delta_t_elem.get('value')) if delta_t_elem is not None else 40.0
                    
                    # Szukaj projekcji z "Total length"
                    for projection in u_factor.findall('Projection'):
                        length_type_elem = projection.find('Length-type')
                        if length_type_elem is not None and 'Total length' in length_type_elem.text:
                            
                            # Pobierz długość
                            length_elem = projection.find('Length')
                            length_value = float(length_elem.get('value')) if length_elem is not None else 0.0
                            
                            # Pobierz wartość U
                            u_factor_elem = projection.find('U-factor')
                            if u_factor_elem is not None:
                                u_value_str = u_factor_elem.get('value')
                                if u_value_str and u_value_str != 'NA':
                                    u_value = float(u_value_str)
                                    
                                    # OBLICZENIE STRUMIENIA CIEPŁA (tak jak w Twoim skrypcie)
                                    heat_flux_per_meter = u_value * delta_t  # W/m
                                    total_heat_flux = heat_flux_per_meter * (length_value / 1000.0)  # zamiana mm na m
                                    
                                    print(f"🔥 Obliczono strumień ciepła dla PHI: {total_heat_flux:.6f} W")
                                    print(f"   U = {u_value}, ΔT = {delta_t}, L = {length_value}mm")
                                    return total_heat_flux
            
            print("❌ Nie znaleziono strumienia ciepła dla PHI w pliku .thmx")
            return None
            
        except Exception as e:
            print(f"❌ Błąd ekstrakcji strumienia ciepła: {e}")
            return None




    def find_temperatures_from_thmx(self, thmx_file):
        """Znajduje temperatury z pliku .thmx"""
        try:
            with open(thmx_file, 'r', encoding='utf-8') as file:
                content = file.read()
            
            content = content.replace(' xmlns="http://windows.lbl.gov"', '')
            root = ET.fromstring(content)
            
            temperatures = {'Ti': None, 'Te': None}
            
            # Szukaj BoundaryConditions
            for bc in root.findall('.//BoundaryCondition'):
                name_elem = bc.find('Name')
                if name_elem is not None:
                    name = name_elem.text
                    temp_elem = bc.find('Temperature')
                    if temp_elem is not None:
                        temp_value = float(temp_elem.get('value'))
                        
                        if 'Ti' in name:
                            temperatures['Ti'] = temp_value
                        elif 'Te' in name:
                            temperatures['Te'] = temp_value
            
            # Alternatywne wyszukiwanie w BoundaryConditions
            if temperatures['Ti'] is None or temperatures['Te'] is None:
                for bc in root.findall('.//BoundaryCondition'):
                    bc_name = bc.get('Name', '')
                    temp_value = float(bc.get('Temperature', 0))
                    
                    if 'Ti' in bc_name and temperatures['Ti'] is None:
                        temperatures['Ti'] = temp_value
                    elif 'Te' in bc_name and temperatures['Te'] is None:
                        temperatures['Te'] = temp_value
            
            print(f"Znalezione temperatury: Ti={temperatures['Ti']}, Te={temperatures['Te']}")
            return temperatures
            
        except Exception as e:
            print(f"Błąd wyszukiwania temperatur: {e}")
            return {'Ti': None, 'Te': None}
    
    def get_u_values_from_selected_curves(self):
        """Pobiera wartości U z Geometry Nodes zaznaczonych krzywych USection"""
        u_values = {}
        
        for obj in bpy.context.selected_objects:
            if obj.type == 'CURVE' and obj.name.startswith('USection_'):
                # Pobierz numer USection (np. USection_U1 -> U1)
                usection_num = obj.name.replace('USection_', '')
                
                # Sprawdź czy obiekt ma Geometry Nodes z wartością U
                for modifier in obj.modifiers:
                    if modifier.type == 'NODES' and modifier.node_group:
                        # Sprawdź dostępne socket-y
                        available_inputs = list(modifier.keys())
                        
                        # Szukaj socketu U-Value
                        u_value_sockets = [s for s in available_inputs if 'u-value' in s.lower() or 'u_value' in s.lower() or s == 'Socket_26']
                        
                        if u_value_sockets:
                            try:
                                u_value = modifier[u_value_sockets[0]]
                                u_values[usection_num] = u_value
                                print(f"Znaleziono U-Value dla {usection_num}: {u_value}")
                                break
                            except Exception as e:
                                print(f"Błąd pobierania U-Value dla {usection_num}: {e}")
        
        return u_values
    
    def get_lengths_from_selected_curves(self):
        """Pobiera długości zaznaczonych krzywych USection"""
        curve_lengths = {}
        
        for obj in bpy.context.selected_objects:
            if obj.type == 'CURVE' and obj.name.startswith('USection_'):
                # Pobierz numer USection
                usection_num = obj.name.replace('USection_', '')
                
                # Oblicz długość krzywej
                if obj.data.splines:
                    spline = obj.data.splines[0]
                    if len(spline.points) >= 2:
                        # Oblicz długość między pierwszym a ostatnim punktem
                        point1 = obj.matrix_world @ spline.points[0].co.xyz
                        point2 = obj.matrix_world @ spline.points[-1].co.xyz
                        length = (point2 - point1).length
                        
                        curve_lengths[usection_num] = length
                        print(f"Długość krzywej {usection_num}: {length:.3f} m")
        
        return curve_lengths
    
    def copy_and_fill_excel_template(self, template_file, output_file, phi_data, temperatures, u_values, curve_lengths, thmx_file, phi_heat_flux=None):
        """Kopiuje i wypełnia szablon Excel - z Ti, Te, strumieniem ciepła PHI i wartościami PHI"""
        try:
            # Spróbuj dodać ścieżkę użytkownika do sys.path
            if not self.add_user_site_packages():
                return self.create_fallback_data_file(output_file, phi_data, temperatures, u_values, curve_lengths, thmx_file, phi_heat_flux)
            
            # Teraz spróbuj zaimportować openpyxl
            try:
                from openpyxl import load_workbook
            except ImportError:
                print("❌ openpyxl wciąż niedostępny")
                return self.create_fallback_data_file(output_file, phi_data, temperatures, u_values, curve_lengths, thmx_file, phi_heat_flux)
            
            # Skopiuj plik wzorca
            shutil.copy2(template_file, output_file)
            print(f"📋 Skopiowano wzorzec do: {output_file}")
            
            # Wczytaj skopiowany plik
            workbook = load_workbook(output_file)
            
            # Sprawdź czy arkusz "Mostki Termiczne" istnieje
            if "Mostki Termiczne" not in workbook.sheetnames:
                print("❌ Błąd: Arkusz 'Mostki Termiczne' nie istnieje!")
                # Spróbuj użyć pierwszego arkusza
                sheet_name = workbook.sheetnames[0]
                print(f"Używam arkusza: {sheet_name}")
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook["Mostki Termiczne"]
                print("📊 Pracuję na arkuszu: 'Mostki Termiczne'")
            
            # MAPOWANIE KOMÓREK
            u_cell_mapping = {
                'U1': 'B30', 'U2': 'D30', 'U3': 'F30', 'U4': 'H30', 'U5': 'J30', 'U6': 'L30',
                'U7': 'B34', 'U8': 'D34', 'U9': 'F34', 'U10': 'H34', 'U11': 'J34', 'U12': 'L34'
            }
            
            length_cell_mapping = {
                'DL1': 'B40', 'DL2': 'D40', 'DL3': 'F40', 'DL4': 'H40', 'DL5': 'J40', 'DL6': 'L40',
                'DL7': 'B44', 'DL8': 'D44', 'DL9': 'F44', 'DL10': 'H44', 'DL11': 'J44', 'DL12': 'L44'
            }
            
            # MAPOWANIE KOMÓREK DLA PHI
            phi_cell_mapping = {
                'PHI1': 'B50', 'PHI2': 'D50', 'PHI3': 'F50', 'PHI4': 'H50', 'PHI5': 'J50', 'PHI6': 'L50',
                'PHI7': 'B54', 'PHI8': 'D54', 'PHI9': 'F54', 'PHI10': 'H54', 'PHI11': 'J54', 'PHI12': 'L54'
            }
            
            # 1. Wpisz temperatury Ti i Te
            if temperatures['Ti'] is not None:
                worksheet['F12'] = temperatures['Ti']
                print(f"✅ Wpisano Ti = {temperatures['Ti']} °C do komórki F12")
            
            if temperatures['Te'] is not None:
                worksheet['H12'] = temperatures['Te']
                print(f"✅ Wpisano Te = {temperatures['Te']} °C do komórki H12")
            
            # 2. Wpisz strumień ciepła PHI do B20
            if phi_heat_flux is not None:
                worksheet['B20'] = phi_heat_flux
                print(f"✅ Wpisano strumień ciepła PHI = {phi_heat_flux:.6f} W do komórki B20")
            else:
                # Spróbuj pobrać strumień ciepła jeśli nie został przekazany
                heat_flux = self.extract_heat_flux_from_thmx(thmx_file)
                if heat_flux is not None:
                    worksheet['B20'] = heat_flux
                    print(f"✅ Wpisano strumień ciepła PHI = {heat_flux:.6f} W do komórki B20")
                else:
                    print("⚠️  Nie znaleziono strumienia ciepła dla PHI - komórka B20 pozostanie pusta")
            
            # 3. Wpisz wartości U z Geometry Nodes
            for usection_num, u_value in u_values.items():
                if usection_num in u_cell_mapping:
                    cell = u_cell_mapping[usection_num]
                    worksheet[cell] = u_value
                    print(f"✅ Wpisano {usection_num} = {u_value:.6f} W/m²K do komórki {cell}")
            
            # 4. Wpisz długości krzywych
            for usection_num, length in curve_lengths.items():
                dl_section = 'DL' + usection_num[1:] if usection_num.startswith('U') else usection_num
                if dl_section in length_cell_mapping:
                    cell = length_cell_mapping[dl_section]
                    worksheet[cell] = length
                    print(f"✅ Wpisano {dl_section} = {length:.3f} m do komórki {cell}")
            
            # 5. Wpisz wartości PHI z .thmx
            for phi_name, phi_value in phi_data.items():
                if phi_name in phi_cell_mapping:
                    cell = phi_cell_mapping[phi_name]
                    worksheet[cell] = phi_value
                    print(f"✅ Wpisano {phi_name} = {phi_value:.6f} do komórki {cell}")
            
            # Zapisz zmiany
            workbook.save(output_file)
            workbook.close()
            
            print(f"🎉 Pomyślnie wypełniono plik Excel!")
            
            # Podsumowanie
            print("\n📋 PODSUMOWANIE WYPEŁNIONYCH DANYCH:")
            print(f"   🌡️  Ti: {temperatures['Ti']} °C (F12)")
            print(f"   🌡️  Te: {temperatures['Te']} °C (H12)")
            if phi_heat_flux is not None:
                print(f"   🔥 Strumień ciepła PHI: {phi_heat_flux:.6f} W (B20)")
            print(f"   🔥 Wartości U: {len(u_values)}")
            print(f"   📏 Długości: {len(curve_lengths)}")
            print(f"   Φ Wartości PHI: {len(phi_data)}")
            
            return True
            
        except Exception as e:
            print(f"❌ Błąd przy wypełnianiu szablonu Excel: {e}")
            import traceback
            traceback.print_exc()
            return self.create_fallback_data_file(output_file, phi_data, temperatures, u_values, curve_lengths, thmx_file, phi_heat_flux)


    def calculate_total_heat_flux(self, u_values, curve_lengths, temperatures):
        """Oblicza całkowity strumień ciepła przez wszystkie mostki termiczne"""
        try:
            if temperatures['Ti'] is None or temperatures['Te'] is None:
                print("❌ Brak temperatur do obliczenia strumienia ciepła")
                return None
            
            deltaT = temperatures['Ti'] - temperatures['Te']
            print(f"🌡️  DeltaT = {temperatures['Ti']} - {temperatures['Te']} = {deltaT} °C")
            
            total_heat_flux = 0.0
            
            print("🔥 OBLICZANIE STRUMIENIA CIEPŁA:")
            print("Mostek | U [W/m²K] | Długość [m] | ΔT [°C] | Strumień [W]")
            print("-------|------------|-------------|----------|-------------")
            
            for usection_num, u_value in u_values.items():
                if usection_num in curve_lengths:
                    length = curve_lengths[usection_num]
                    section_heat_flux = u_value * deltaT * length
                    total_heat_flux += section_heat_flux
                    
                    print(f"{usection_num:6} | {u_value:10.6f} | {length:11.3f} | {deltaT:8} | {section_heat_flux:11.6f}")
                else:
                    print(f"{usection_num:6} | {u_value:10.6f} | {'BRAK':11} | {deltaT:8} | {'BRAK':11}")
            
            print(f"SUMA STRUMIENIA CIEPŁA: {total_heat_flux:.6f} W")
            return total_heat_flux
            
        except Exception as e:
            print(f"❌ Błąd obliczania strumienia ciepła: {e}")
            return None

    def add_user_site_packages(self):
        """Dodaje ścieżkę do site-packages użytkownika do sys.path"""
        import sys
        import os
        
        # Ścieżka do site-packages użytkownika
        user_site_packages = os.path.join(os.path.expanduser("~"), 
                                        "AppData", "Roaming", 
                                        "Python", "Python311", 
                                        "site-packages")
        
        if os.path.exists(user_site_packages) and user_site_packages not in sys.path:
            sys.path.append(user_site_packages)
            print(f"✅ Dodano ścieżkę użytkownika: {user_site_packages}")
            return True
        
        print(f"❌ Nie znaleziono ścieżki: {user_site_packages}")
        return True

    def create_fallback_data_file(self, output_file, phi_data, temperatures, u_values, curve_lengths, thmx_file, phi_heat_flux=None):
        """Tworzy plik tekstowy z danymi gdy Excel nie jest dostępny - zaktualizowana z PHI"""
        try:
            data_file = output_file.replace('.xlsx', '_DANE.txt')
            
            # Pobierz strumień ciepła z .thmx jeśli nie został przekazany
            if phi_heat_flux is None:
                phi_heat_flux = self.extract_heat_flux_from_thmx(thmx_file)
            
            with open(data_file, 'w', encoding='utf-8') as f:
                f.write("DANE EKSPORTOWANE Z BLENDERA\n")
                f.write("=" * 50 + "\n")
                f.write("Plik Excel nie mógł zostać wypełniony (brak openpyxl)\n")
                f.write("Skopiuj poniższe wartości ręcznie do pliku wzorca.xlsx\n")
                f.write("=" * 50 + "\n\n")
                
                f.write("🌡️  TEMPERATURY (wpisz ręcznie do Excel):\n")
                f.write("F12 - Ti: {} °C\n".format(temperatures['Ti'] if temperatures['Ti'] is not None else "BRAK"))
                f.write("H12 - Te: {} °C\n".format(temperatures['Te'] if temperatures['Te'] is not None else "BRAK"))
                
                if phi_heat_flux is not None:
                    f.write("B20 - Strumień ciepła PHI: {:.6f} W\n".format(phi_heat_flux))
                else:
                    f.write("B20 - Strumień ciepła: NIE ZNALEZIONO W .thmx\n")
                
                f.write("\n" + "=" * 50 + "\n")
                f.write("🗺️  MAPOWANIE KOMÓREK DO EXCEL:\n")
                f.write("=" * 50 + "\n")
                
                f.write("\n🔥 WARTOŚCI U (komórki B30-L34):\n")
                u_mapping = {
                    'U1': 'B30', 'U2': 'D30', 'U3': 'F30', 'U4': 'H30', 'U5': 'J30', 'U6': 'L30',
                    'U7': 'B34', 'U8': 'D34', 'U9': 'F34', 'U10': 'H34', 'U11': 'J34', 'U12': 'L34'
                }
                
                for usection_num, u_value in u_values.items():
                    if usection_num in u_mapping:
                        f.write("{} - {}: {:.6f} W/m²K\n".format(u_mapping[usection_num], usection_num, u_value))
                
                f.write("\n📏 DŁUGOŚCI (komórki B40-L44):\n")
                length_mapping = {
                    'DL1': 'B40', 'DL2': 'D40', 'DL3': 'F40', 'DL4': 'H40', 'DL5': 'J40', 'DL6': 'L40',
                    'DL7': 'B44', 'DL8': 'D44', 'DL9': 'F44', 'DL10': 'H44', 'DL11': 'J44', 'DL12': 'L44'
                }
                
                for usection_num, length in curve_lengths.items():
                    dl_section = 'DL' + usection_num[1:] if usection_num.startswith('U') else usection_num
                    if dl_section in length_mapping:
                        f.write("{} - {}: {:.3f} m\n".format(length_mapping[dl_section], dl_section, length))
                
                f.write("\nΦ WARTOŚCI PHI (komórki B50-L54):\n")
                phi_mapping = {
                    'PHI1': 'B50', 'PHI2': 'D50', 'PHI3': 'F50', 'PHI4': 'H50', 'PHI5': 'J50', 'PHI6': 'L50',
                    'PHI7': 'B54', 'PHI8': 'D54', 'PHI9': 'F54', 'PHI10': 'H54', 'PHI11': 'J54', 'PHI12': 'L54'
                }
                
                for phi_name, phi_value in phi_data.items():
                    if phi_name in phi_mapping:
                        f.write("{} - {}: {:.6f}\n".format(phi_mapping[phi_name], phi_name, phi_value))
            
            print(f"📄 Utworzono plik z instrukcjami: {data_file}")
            print("ℹ️  Skopiuj wartości ręcznie do pliku wzorca.xlsx")
            
            return True
            
        except Exception as e:
            print(f"❌ Błąd przy tworzeniu pliku z danymi: {e}")
            return False
# Lista wszystkich klas operatorów do rejestracji
classes = (
    THERM_OT_check_normals,
    THERM_OT_check_vertices,
    THERM_OT_round_vertices,
    THERM_OT_force_round_vertices,
    THERM_OT_recalc_normals_confirm,
    THERM_OT_round_vertices_confirm,
    THERM_OT_flip_downward_faces,
    THERM_OT_quick_flip_downward_faces,
    THERM_OT_create_ti_edges,
    THERM_OT_create_te_edges,
    THERM_OT_create_adiabatic_edges,
    THERM_OT_create_adiabatic_auto,
    THERM_OT_create_ti_auto,
    THERM_OT_create_te_auto,
    THERM_OT_create_ufactor_auto,
    THERM_OT_create_ufactor_edges,
    THERM_OT_export_to_therm,
    THERM_OT_run_therm_calculation_thmx,
    THERM_OT_run_therm_calculation_thm,
    THERM_OT_open_therm_folder,
    THERM_OT_import_from_therm,
    THERM_OT_clean_to_boundary,
    THERM_OT_create_usection_1,
    THERM_OT_create_usection_2,
    THERM_OT_create_usection_3,
    THERM_OT_create_usection_4,
    THERM_OT_create_usection_5,
    THERM_OT_create_usection_6,
    THERM_OT_create_usection_7,
    THERM_OT_create_usection_8,
    THERM_OT_create_usection_9,
    THERM_OT_create_usection_10,
    THERM_OT_create_usection_11,
    THERM_OT_create_usection_12,
    THERM_OT_debug_usections,
    THERM_OT_debug_sockets,
    THERM_OT_export_to_excel  # DODAJ TĘ LINIĘ
)

def register():
    for cls in classes:
        try:
            bpy.utils.register_class(cls)
        except Exception as e:
            print(f"Błąd rejestracji {cls}: {e}")

def unregister():
    for cls in reversed(classes):
        try:
            bpy.utils.unregister_class(cls)
        except Exception as e:
            print(f"Błąd wyrejestrowania {cls}: {e}")